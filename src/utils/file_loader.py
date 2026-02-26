from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders.pdf import PyPDFLoader
from langchain_community.document_loaders.word_document import Docx2txtLoader
from src.utils.text_sanitizer import sanitize_text
from openpyxl import load_workbook
from datetime import date, datetime
import csv


def read_csv(file_path: str) -> tuple[list[str], int]:
    """
    读取CSV文件并返回切分后的文本块列表和块数。
    这里的实现非常简单，直接将每行作为一个文本块。根据需要可以进行更复杂的处理。
    """
    chunks = []
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            line = " ".join(row)  # 将每行的所有列合并成一个字符串
            chunks.append(sanitize_text(line))

    print(f"总共切分成 {len(chunks)} 个文本块")
    return chunks, len(chunks)


def _find_column_index(fields: list[str], candidates: list[str]) -> int | None:
    lowered = [field.lower() for field in fields]
    for candidate in candidates:
        if candidate in fields:
            return fields.index(candidate)
        candidate_lower = candidate.lower()
        if candidate_lower in lowered:
            return lowered.index(candidate_lower)
    return None


def _normalize_optional_source(value: object) -> str | None:
    if value is None:
        return None
    source = sanitize_text(str(value).strip())
    return source or None


def _normalize_optional_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text_value = str(value).strip()
    if not text_value:
        return None

    try:
        return datetime.fromisoformat(text_value).date()
    except ValueError:
        pass

    for fmt in ("%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue

    raise ValueError(
        f"Invalid 日期/date value: {text_value}. "
        "Supported formats: YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD, YYYYMMDD, YYYY年MM月DD日"
    )


def read_qa_excel(file_path: str) -> list[dict[str, str | date | None]]:
    """
    读取“问题/答案”两列的Excel（.xlsx/.xlsm，默认首个sheet）。
    返回: [{"question": "...", "answer": "...", "source": "...", "date": date|None}, ...]
    其中 source/date 列可选（支持中文列名“来源/日期”或英文“source/date”）。
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel has no active worksheet.")
        row_iter = ws.iter_rows(values_only=True)
        header = next(row_iter, None)
        if header is None:
            raise ValueError("Excel is empty or missing header row.")

        fields = [str(col).strip() if col is not None else "" for col in header]
        question_idx = _find_column_index(fields, ["问题", "question"])
        answer_idx = _find_column_index(fields, ["答案", "answer"])
        source_idx = _find_column_index(fields, ["来源", "source"])
        date_idx = _find_column_index(fields, ["日期", "date"])

        if question_idx is None or answer_idx is None:
            raise ValueError("Excel must contain columns: 问题/答案 (or question/answer)")

        records: list[dict[str, str | date | None]] = []
        for row_no, row in enumerate(row_iter, start=2):
            question_value = row[question_idx] if question_idx < len(row) else None
            answer_value = row[answer_idx] if answer_idx < len(row) else None
            source_value = row[source_idx] if source_idx is not None and source_idx < len(row) else None
            date_value = row[date_idx] if date_idx is not None and date_idx < len(row) else None

            question = sanitize_text(str(question_value).strip()) if question_value is not None else ""
            answer = sanitize_text(str(answer_value).strip()) if answer_value is not None else ""
            if not question or not answer:
                continue
            try:
                info_date = _normalize_optional_date(date_value)
            except ValueError as e:
                raise ValueError(f"Row {row_no}: {e}") from e
            records.append(
                {
                    "question": question,
                    "answer": answer,
                    "source": _normalize_optional_source(source_value),
                    "date": info_date,
                }
            )

        return records
    finally:
        wb.close()


def read_excel(file_path: str) -> tuple[list[str], int]:
    """
    读取“问题/答案”两列Excel，并将每行转为一个文本块。
    """
    records = read_qa_excel(file_path)
    chunks = [
        sanitize_text(f"问题：{record['question']}\n答案：{record['answer']}")
        for record in records
    ]
    print(f"总共切分成 {len(chunks)} 个文本块")
    return chunks, len(chunks)


def read_txt(file_path: str) -> tuple[list[str], int]:
    """
    读取TXT文件并返回切分后的文本块列表和块数。
    """
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=512,
        chunk_overlap=200,
        length_function=len,
    )

    text = ""
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                text = f.read()
            break
        except UnicodeDecodeError as e:
            last_error = e
            continue
    else:
        raise ValueError(
            "Unsupported TXT encoding. Please use UTF-8/UTF-8-SIG/GB18030."
        ) from last_error

    text = sanitize_text(text)
    chunks = text_splitter.split_text(text)
    chunks = [sanitize_text(chunk) for chunk in chunks]

    print(f"总共切分成 {len(chunks)} 个文本块")
    return chunks, len(chunks)


def read_pdf(file_path: str) -> tuple[list[str], int]:
    """
    读取PDF文件并返回切分后的文本块列表和块数。
    """
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=512, chunk_overlap=200, length_function=len)
    loader = PyPDFLoader(file_path)

    texts = [doc.page_content for doc in loader.lazy_load()]
    texts = "".join(texts)
    texts = sanitize_text(texts)
    
    chunks = text_splitter.split_text(texts)
    chunks = [sanitize_text(chunk) for chunk in chunks]

    print(f"总共切分成 {len(chunks)} 个文本块")
    return chunks, len(chunks)


def read_docx(file_path: str) -> tuple[list[str], int]:
    """
    读取Word文档并返回切分后的文本块列表和块数。
    """
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=512, chunk_overlap=200, length_function=len)
    loader = Docx2txtLoader(file_path)

    texts = [doc.page_content for doc in loader.lazy_load()]
    texts = "".join(texts)
    texts = sanitize_text(texts)
    
    chunks = text_splitter.split_text(texts)
    chunks = [sanitize_text(chunk) for chunk in chunks]

    print(f"总共切分成 {len(chunks)} 个文本块")
    return chunks, len(chunks)


if __name__ == "__main__":
    pass
    chunks, count = read_pdf("~/Downloads/20250916太原理工大学2025版学生手册（封面＋正文）.pdf")
    print(f"返回了 {count} 个文本块")
    # import json
    # print(json.dumps(texts, ensure_ascii=False, indent=2))
